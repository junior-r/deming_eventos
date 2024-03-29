import datetime
import os.path

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q, QuerySet
from django.db.models.fields.files import ImageFieldFile
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.translation import to_language

from Apps.Eventos.forms import ParticipantDataUpdateForm
from Apps.Eventos.models import Event, Participant
from Apps.Users.forms import UserForm, UpdateUserForm
from Apps.Users.models import User


def sign_up(request):
    data = {
        'form': UserForm(),
    }

    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES)
        if form.is_valid():
            password1 = request.POST.get('password1')
            profile_image_user = request.FILES.get('profile_image')
            curriculum = request.FILES.get('curriculum')
            username = request.POST.get('username')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            profession = request.POST.get('profession')
            interests = request.POST.getlist('interests')

            user = User(
                first_name=first_name, last_name=last_name, email=email, username=username, password=password1,
                profile_image_user=profile_image_user, curriculum=curriculum, profession=profession
            )
            user.save()
            user.interests.add(*interests)
            login(request, user)

            messages.success(request, 'Cuenta creada exitosamente.')
            return redirect('index')
        else:
            messages.error(request, 'Algún dato es inválido. Intente de nuevo!')
            data['form'] = form

    return render(request, 'registration/sign_up.html', data)


def password_reset(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        if email is not None and email != '':
            user = get_object_or_404(User, email=email)
            subject = 'Instrucciones para cambiar la contraseña'
            message = f'¡Hola! @{user.username.capitalize()} \nResives este correo por parte de la plataforma de ' \
                      f'eventos de Instituto Superior Tecnológico Corporativo Edwards Deming \n\n' \
                      f'Nos contactamos contigo ya que solicitaste cambiar tu contraseña para ingresar a la ' \
                      f'plataforma \n\nSolo ve al siguiente link y podrás cambiarla {request.get_host()}/accounts' \
                      f'/reset_password_form/{user.id}/ \n\n¡Que tengas buen día! \nDeming Team'
            try:
                user.email_user(subject=subject, message=message, from_email=settings.EMAIL_HOST_USER)
                messages.success(request, 'Correo enviado exitosamente')
                return redirect('custom_reset_password_done', user.id)
            except Exception as e:
                messages.error(request, f'Error: {e}')
    return render(request, 'registration/password_reset.html')


def reset_password_done(request, user_id):
    user = get_object_or_404(User, id=user_id)
    data = {
        'user': user,
    }
    return render(request, 'registration/custom_password_reset_done.html', data)


def reset_password_form(request, user_id):
    user = get_object_or_404(User, id=user_id)
    data = {
        'user': user,
    }

    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        if password1 and password2 and password1 == password2:
            if user.check_password(password1):
                messages.error(request, 'No puedes usar la contraseña anterior.')
                return redirect('custom_reset_password_form', user.id)
            else:
                user.password = password1
                user.set_password(password1)
                user.save()
                login(request, user)
                messages.success(request, '¡Contraseña cambiada exitosamente. Bienvenido!')
                return redirect('index')
        else:
            messages.error(request, 'Debes prpoporcionar 1 contraseña 2 veces, estas deben ser iguales. Tome en cuenta que se distingue entre mayúsculas y minúsculas')
            return redirect('custom_reset_password_form', user.id)

    return render(request, 'registration/custom_reset_password_form.html', data)


@login_required
def profile(request, username, id_user):
    user = get_object_or_404(User, id=id_user, username=username)
    permissions_content_type = get_object_or_404(ContentType, app_label='auth', model='permission')
    users_content_type = get_object_or_404(ContentType, app_label='Users', model='user')
    careers_content_type = get_object_or_404(ContentType, app_label='Eventos', model='career')
    event_content_type = get_object_or_404(ContentType, app_label='Eventos', model='event')
    event_participant_content_type = get_object_or_404(ContentType, app_label='Eventos', model='eventparticipant')
    participant_content_type = get_object_or_404(ContentType, app_label='Eventos', model='participant')

    permissions = Permission.objects.filter(
        Q(content_type=permissions_content_type) | Q(content_type=users_content_type) |
        Q(content_type=careers_content_type) | Q(content_type=event_content_type) |
        Q(content_type=event_participant_content_type) | Q(content_type=participant_content_type)
    )

    current_user = get_object_or_404(User, id=request.user.id)
    current_user_perms = Permission.objects.filter(user=user)
    events_teacher = None
    if user.is_teacher:
        events_teacher = Event.objects.filter(teachers__username=user.username)

    data = {
        'current_profile_user': user,
        'events_teacher': events_teacher,
        'current_user': current_user,
        'current_user_perms': current_user_perms,
        'permissions': permissions,
        'form': UpdateUserForm(instance=user)
    }

    is_staff_original = user.is_staff
    is_active_original = user.is_active
    is_teacher_original = user.is_teacher
    is_referral_original = user.is_referral

    if request.method == 'POST':
        if request.user.id != user.id:
            if not request.user.is_superuser:
                messages.error(request, 'No tienes permiso para actualizar esta información')
                return redirect('profile', user.username, user.id)
        else:
            pass

        form = UpdateUserForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            picture_profile = request.FILES.get('profile_image') if request.FILES.get('profile_image') is not None \
                else user.profile_image_user.name
            curriculum = request.FILES.get('curriculum') if request.FILES.get('curriculum') is not None \
                else user.curriculum.name
            username = request.POST.get('username') if request.POST.get('username') is not None else \
                user.username
            first_name = request.POST.get('first_name') if request.POST.get('first_name') is not None else \
                user.first_name
            last_name = request.POST.get('last_name') if request.POST.get('last_name') is not None else \
                user.last_name
            profession = request.POST.get('profession') if request.POST.get('profession') is not None else \
                user.profession
            email = request.POST.get('email') if request.POST.get('email') is not None else \
                user.email
            is_staff_form = request.POST.get('is_staff')
            is_active_form = request.POST.get('is_active')
            is_teacher_form = request.POST.get('is_teacher')
            is_referral_form = request.POST.get('is_referral')
            if is_staff_form is None or is_staff_form == 'off':
                is_staff_form = False
            else:
                is_staff_form = True

            if is_active_form is None or is_active_form == 'on':
                is_active_form = True
            else:
                is_active_form = False

            if is_teacher_form == 'on':
                is_teacher_form = True
            else:
                is_teacher_form = False

            if is_referral_form == 'on':
                is_referral_form = True
            else:
                is_referral_form = False

            user.profile_image_user = picture_profile
            user.curriculum = curriculum
            user.username = username
            user.profession = profession
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            new_interests = request.POST.getlist('interests')

            if user.id == current_user.id:
                user.is_staff = is_staff_original
                user.is_active = is_active_original
                user.is_teacher = is_teacher_original
                user.is_referral = is_referral_original
            else:
                user.is_staff = is_staff_form
                user.is_active = is_active_form
                user.is_teacher = is_teacher_form
                user.is_referral = is_referral_form

            user.interests.set(new_interests)
            user.save()
            user.refresh_from_db()
            if not is_teacher_form:
                event = Event.objects.filter(teachers__username=user.username)
                for e in event.all():
                    e.teachers.remove(user)

            messages.success(request, 'La información ha sido actualizada')
            return redirect('profile', user.username, user.id)
        else:
            data['form'] = form
            messages.error(request, 'Datos inválidos')
    return render(request, 'Users/profile.html', data)


@login_required
def info_participant(request, username, id_user):
    current_user = get_object_or_404(User, username=username, id=id_user)
    participant = get_object_or_404(Participant, user_id=current_user.id)
    img_participant = participant.get_profile_image()

    data = {
        'current_user': current_user,
        'participant': participant,
        'form': ParticipantDataUpdateForm(instance=participant)
    }

    if request.method == 'POST':
        form = ParticipantDataUpdateForm(request.POST, request.FILES, instance=participant)
        if form.is_valid():
            if form.has_changed():
                profile_image = request.FILES.get('profile_image') if request.FILES.get('profile_image') is not None \
                    else img_participant

                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                country_of_birth = request.POST.get('country_of_birth')
                dni = request.POST.get('dni')
                passport_number = request.POST.get('passport_number') if request.POST.get('passport_number') else None
                gender = request.POST.get('gender')
                birthdate = request.POST.get('birthdate')
                current_country = request.POST.get('current_country')
                address = request.POST.get('address')
                phone = request.POST.get('phone')
                email = request.POST.get('email')
                alternative_email = request.POST.get('alternative_email') if request.POST.get(
                    'alternative_email') else None
                object = request.POST.get('object')

                participant.profile_image, participant.first_name = profile_image, first_name
                participant.last_name, participant.country_of_birth = last_name, country_of_birth
                participant.dni, participant.passport_number = dni, passport_number
                participant.gender, participant.birthdate = gender, birthdate
                participant.current_country, participant.address = current_country, address
                participant.phone, participant.email, participant.alternative_email = phone, email, alternative_email
                participant.object = object
                participant.save()

                messages.success(request, 'Información actualizada exitósamente')
                return redirect('info_participant', current_user.username, current_user.id)
            else:
                messages.info(request, 'Debes cambiar algún campo para actualizar tus datos.')
                return redirect('info_participant', current_user.username, current_user.id)
        else:
            messages.error(request, 'Ocurrió un error. Revisa e intenta de nuevo.')
            data['form'] = form

    return render(request, 'Users/info_participant.html', data)


@login_required
def user_events(request, username, id_user):
    current_user = get_object_or_404(User, username=username, id=id_user)
    if current_user.is_teacher:
        events_teacher = Event.objects.filter(teachers__username=current_user.username)
        # Paginate teachers
        if events_teacher:
            paginator_teachers = Paginator(events_teacher, 9)
            page_number_teachers = request.GET.get('page')
            events_teacher_data = paginator_teachers.get_page(page_number_teachers)
        try:
            participant = get_object_or_404(Participant, user_id=current_user.id)
            events = Event.objects.filter(eventparticipant__participant__id=participant.id, active=True)
            # Paginate participants
            if events:
                paginator_participants = Paginator(events, 9)
                page_number_participants = request.GET.get('page')
                events_participant_data = paginator_participants.get_page(page_number_participants)
        except:
            participant = None
            events = None
            events_participant_data = None
    else:
        events_teacher = None
        events_teacher_data = None
        participant = get_object_or_404(Participant, user_id=current_user.id)
        events = Event.objects.filter(eventparticipant__participant__id=participant.id, eventparticipant__pay=True)
        # Paginate participants
        paginator_participants = Paginator(events, 9)
        page_number_participants = request.GET.get('page')
        events_participant_data = paginator_participants.get_page(page_number_participants)

    data = {
        'user': current_user,
        'participant': participant,
        'events': events_participant_data,
        'events_teacher': events_teacher_data,
    }

    if request.method == 'POST':
        query_participants = request.POST.get('query-events-participant')
        query_teachers = request.POST.get('query-events-teachers')
        # Find participant events
        if query_participants is not None:
            filtered_events_participant = events.filter(
                Q(title__icontains=query_participants))
            if filtered_events_participant:
                # Paginate Filtered participants
                paginator_filtered_participants = Paginator(filtered_events_participant, 9)
                page_number_filtered_participants = request.GET.get('page')
                events_filtered_participant_data = paginator_filtered_participants.get_page(page_number_filtered_participants)

                data['filtered_events_participant'] = events_filtered_participant_data
                data['filtered_events_participant_count'] = filtered_events_participant.all().count()
                data['query_participants'] = query_participants
            else:
                messages.info(request, 'No se encontraron coincidencias')
        # Find teacher events
        elif query_teachers is not None:
            filtered_events_teachers = events_teacher.filter(
                Q(title__icontains=query_teachers))
            if filtered_events_teachers:
                # Paginate Filtered teachers
                paginator_filtered_teachers = Paginator(filtered_events_teachers, 9)
                page_number_filtered_teachers = request.GET.get('page')
                events_filtered_teacher_data = paginator_filtered_teachers.get_page(page_number_filtered_teachers)

                data['filtered_events_teachers'] = events_filtered_teacher_data
                data['filtered_events_teachers_count'] = filtered_events_teachers.all().count()
                data['query_teachers'] = query_teachers
            else:
                messages.info(request, 'No se encontraron coincidencias')
        else:
            messages.info(request, 'Escríbe algo para buscar...')
            return redirect('user_events', request.user.username, request.user.id)

    return render(request, 'Users/user_events.html', data)


@login_required
@permission_required('auth.add_permission', raise_exception=True)
@user_passes_test(lambda u: u.is_superuser)
def assign_perms_user(request, current_profile_user_id):
    current_profile_user = get_object_or_404(User, id=current_profile_user_id)
    perms_to_set = request.POST.getlist('permisos')

    try:
        current_profile_user.user_permissions.set(perms_to_set)
        messages.success(request, 'Permisos configurados exitosamente')
    except:
        messages.error(request, 'No se pudo asignar los permisos seleccionado')
    finally:
        return redirect('profile', current_profile_user.username, current_profile_user.id)


@login_required
@permission_required('Users.view_user', raise_exception=True)
def users(request):
    user = get_object_or_404(User, id=request.user.id)
    staff_users = User.objects.filter(is_staff=True).exclude(id=user.id)
    teachers = User.objects.filter(is_teacher=True).exclude(id=user.id)
    referral_users = User.objects.filter(is_referral=True).exclude(id=user.id)
    mortal_users = User.objects.filter(is_staff=False, is_teacher=False, is_referral=False, is_superuser=False)
    data = {
        'staff_users': staff_users,
        'teachers': teachers,
        'referral_users': referral_users,
        'mortal_users': mortal_users,
        'create_users_form': UserForm(),
    }

    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES)
        user_type = request.POST.get('user_type')
        response = create_user(request, user_type, form)
        if response == 'success':
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('users')
        else:
            data['create_users_form'] = form
            messages.error(request, f'Ocurrió un error.')

    return render(request, 'Users/users.html', data)


def export_to_excel(queryset: QuerySet, filename: str):
    # Create an Excel file.
    file = openpyxl.Workbook()
    sheet = file.active

    # Write the column headers to the first row.
    headers = [field.name for field in queryset.model._meta.fields]
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=i + 1).value = header

    # Write the data from the records to the queryset by row.
    for i, record in enumerate(queryset):
        for j, field in enumerate(headers):
            value = getattr(record, field)
            cell = sheet.cell(row=i + 2, column=j + 1)
            # If the value is a date, the value will be formatted as a valid Excel date
            if isinstance(value, datetime.datetime):
                cell.value = f'{value.year}/{value.month}/{value.day} - {value.hour}:{value.minute}:{value.second}'
            elif isinstance(value, ImageFieldFile):
                try:
                    cell.value = str(value.url)
                except ValueError:
                    cell.value = os.path.join(settings.MEDIA_URL, 'user_profile_placeholder.jpg')
                except:
                    cell.value = os.path.join(settings.MEDIA_URL, 'user_profile_placeholder.jpg')
            # if the value is an argon-encoded password, the value will be 'CONFIDENTIAL'
            elif str(value).startswith('argon'):
                cell.value = 'CONDENTIAL'
            else:
                cell.value = str(value)
    # Guarda el archivo de Excel
    file.save("{0}".format(filename))
    return file


@login_required
@permission_required('Users.view_user', raise_exception=True)
def export_users(request, user_type: str, event_id: int):
    current_user = get_object_or_404(User, id=request.user.id)
    users_to_export = None
    if user_type == 'staff':
        users_to_export = User.objects.filter(is_staff=True).exclude(id=current_user.id)
        filename = 'Usuarios_Staff.xlsx'
    elif user_type == 'teacher':
        users_to_export = User.objects.filter(is_teacher=True).exclude(id=current_user.id)
        filename = 'Usuarios_Profesores.xlsx'
    elif user_type == 'referral':
        users_to_export = User.objects.filter(is_referral=True).exclude(id=current_user.id)
        filename = 'Usuarios_Recomendadores.xlsx'
    elif user_type == 'actives_participants':
        event = get_object_or_404(Event, id=event_id)
        participants = Participant.objects.filter(event__participants__event=event,
                                                  event__eventparticipant__active=True)
        users_to_export = participants
        filename = 'Participantes_Activos_{0}.xlsx'.format(event.get_unicode())
    elif user_type == 'participants_payed':
        event = get_object_or_404(Event, id=event_id)
        participants = Participant.objects.filter(event__participants__event=event, event__eventparticipant__pay=True)
        users_to_export = participants
        filename = 'Participantes_Pagos_{0}.xlsx'.format(event.get_unicode())
    else:
        users_to_export = User.objects.filter(is_staff=False, is_teacher=False, is_referral=False, is_superuser=False)
        filename = 'Usuarios_Normales.xlsx'

    if users_to_export.exists():
        file = export_to_excel(users_to_export, filename)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
        response.write(open('{0}'.format(filename), 'rb').read())
        messages.success(request, 'Usuarios descargados exitosamente')
        file_saved = os.path.join(settings.BASE_DIR, filename)
        os.remove(file_saved)
        return response
    else:
        messages.info(request, 'No se encontraron registros para descargar')
        return redirect(request.META.get('HTTP_REFERER'))


@login_required
def referral_users(request, user_id):
    events = Event.objects.filter(active=True).order_by('start_date')
    referral = get_object_or_404(User, id=user_id, is_referral=True)
    events_participants = []
    count_participants = 0
    for event in events:
        participants = Participant.objects.filter(event=event, referral=referral).order_by('event__title')
        count_participants += participants.count()
        events_participants.append({f'event': event, 'participants': participants})
    users = Participant.objects.filter(referral=referral).order_by('event')
    data = {
        'users': users,
        'events_participants': events_participants,
        'count_participants': count_participants,
    }
    return render(request, 'Users/referral_users.html', data)


@login_required
@permission_required('Users.add_user', raise_exception=True)
def create_user(request, user_type: str, form):
    if request.method == 'POST':
        if form.is_valid():
            user = form.save(request)
            if user_type == 'staff':
                if request.user.is_superuser:
                    user.is_staff = True
                else:
                    raise PermissionDenied()
            elif user_type == 'teacher':
                user.is_teacher = True
            elif user_type == 'referral':
                user.is_referral = True

            user.is_active = True
            user.save()
            return 'success'
        else:
            return False


@login_required
def delete_user(request, id_user):
    user = get_object_or_404(User, id=id_user)
    user.delete()
    messages.success(request, '¡Cuenta eliminada exitosamente!')
    return redirect('index')
